"""fantacalcio.it season statistics workbook provider.

The official "Statistiche" Excel is a separate file from the listone
quotations. This importer never creates Player rows — identity comes from
the listone (via PlayerExternalId). Unmatched rows are skipped and
counted so a missing mapping stays visible rather than spawning a
duplicate player (see AGENTS.md > Player Identity).
"""

from dataclasses import dataclass
from decimal import Decimal, InvalidOperation

import openpyxl

from players.models import PlayerExternalId
from providers.players.fantacalcio_it import PROVIDER, ROLE_TO_POSITION
from stats.services import upsert_player_season_stats

HEADER_SCAN_ROWS = 10

COLUMN_ALIASES = {
    'external_id': {'id'},
    'role': {'r'},
    'name': {'nome'},
    'team': {'squadra'},
    'appearances': {'pv', 'pg'},
    'average_rating': {'mv'},
    'fantasy_average': {'fm', 'mf'},
    'goals': {'gf'},
    'assists': {'ass', 'asf'},
    'yellow_cards': {'amm'},
    'red_cards': {'esp'},
    'own_goals': {'au'},
}


@dataclass
class StatsRow:
    external_id: str
    position: str
    name: str
    team_name: str
    appearances: int = 0
    goals: int = 0
    assists: int = 0
    yellow_cards: int = 0
    red_cards: int = 0
    own_goals: int = 0
    average_rating: Decimal | None = None
    fantasy_average: Decimal | None = None


def parse_stats_workbook(path: str) -> list[StatsRow]:
    """Read the 'Tutti' sheet when present, otherwise the first sheet.

    Position-specific sheets (Portieri, etc.) are ignored so the same
    player is not imported twice from one workbook.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet_name = 'Tutti' if 'Tutti' in wb.sheetnames else wb.sheetnames[0]
    return _parse_sheet(wb[sheet_name])


def _parse_sheet(ws) -> list[StatsRow]:
    header_row_index, headers = _find_header_row(ws)
    if headers is None:
        return []

    index = _column_index(headers)
    required = ('external_id', 'role', 'name', 'team')
    if any(key not in index for key in required):
        return []

    rows: list[StatsRow] = []

    for record in ws.iter_rows(min_row=header_row_index + 1, values_only=True):
        role = _cell(record, index['role'][0])
        name = _cell(record, index['name'][0])
        team_name = _cell(record, index['team'][0])
        external_id = _cell(record, index['external_id'][0])
        if not name or not team_name or role not in ROLE_TO_POSITION:
            continue
        rows.append(StatsRow(
            external_id=str(external_id),
            position=ROLE_TO_POSITION[role],
            name=str(name).strip(),
            team_name=str(team_name).strip(),
            appearances=_as_int(_cell(record, _first(index, 'appearances'))),
            goals=_as_int(_cell(record, _first(index, 'goals'))),
            assists=_as_int(_cell(record, _first(index, 'assists'))),
            yellow_cards=_as_int(_cell(record, _first(index, 'yellow_cards'))),
            red_cards=_as_int(_cell(record, _first(index, 'red_cards'))),
            own_goals=_as_int(_cell(record, _first(index, 'own_goals'))),
            average_rating=_as_decimal(_cell(record, _first(index, 'average_rating'))),
            fantasy_average=_as_decimal(_cell(record, _first(index, 'fantasy_average'))),
        ))
    return rows


def _find_header_row(ws):
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=HEADER_SCAN_ROWS, values_only=True), 1):
        headers = [str(c).strip() if c is not None else '' for c in row]
        normalized = {h.lower() for h in headers}
        if 'id' in normalized and 'nome' in normalized:
            return i, headers
    return None, None


def _column_index(headers: list[str]) -> dict[str, list[int]]:
    index: dict[str, list[int]] = {}
    for i, header in enumerate(headers):
        key = header.strip().lower()
        for field, aliases in COLUMN_ALIASES.items():
            if key in aliases:
                index.setdefault(field, []).append(i)
    return index


def _first(index: dict[str, list[int]], field: str) -> int | None:
    values = index.get(field)
    return values[0] if values else None


def _cell(record, column_index: int | None):
    if column_index is None or column_index >= len(record):
        return None
    return record[column_index]


def _as_int(value) -> int:
    if value is None or value == '':
        return 0
    try:
        return int(value)
    except (TypeError, ValueError):
        try:
            return int(float(value))
        except (TypeError, ValueError):
            return 0


def _as_decimal(value) -> Decimal | None:
    if value is None or value == '':
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        return None


@dataclass
class StatsImportSummary:
    created: int = 0
    updated: int = 0
    unmatched: int = 0


def import_season_stats(rows: list[StatsRow], season) -> StatsImportSummary:
    """Upsert PlayerSeasonStats for players already mapped to fantacalcio.it."""
    from players.models import Team

    summary = StatsImportSummary()
    team_cache: dict[str, Team] = {}

    mappings = {
        m.external_id: m
        for m in PlayerExternalId.objects.select_related('player').filter(
            provider=PROVIDER,
            external_id__in=[row.external_id for row in rows],
        )
    }

    for row in rows:
        mapping = mappings.get(row.external_id)
        if mapping is None:
            summary.unmatched += 1
            continue

        if row.team_name not in team_cache:
            team_cache[row.team_name], _ = Team.objects.get_or_create(name=row.team_name)
        club = team_cache[row.team_name]

        _, created = upsert_player_season_stats(
            player=mapping.player,
            season=season,
            club=club,
            position=row.position,
            appearances=row.appearances,
            goals=row.goals,
            assists=row.assists,
            yellow_cards=row.yellow_cards,
            red_cards=row.red_cards,
            own_goals=row.own_goals,
            average_rating=row.average_rating,
            fantasy_average=row.fantasy_average,
        )
        if created:
            summary.created += 1
        else:
            summary.updated += 1

    return summary
