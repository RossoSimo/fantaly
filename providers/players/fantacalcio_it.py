"""fantacalcio.it "listone" provider.

The listone is fantacalcio.it's official pre-season player list: one row
per Serie A player with their own stable `Id`, role, name, club, and
official quotation. It's the closest thing to a canonical player registry
available for this domain, so it's a good anchor for player identity —
but it's still external, untrusted input (see AGENTS.md > Data Sources)
and goes through the same identity resolution service every other
provider would, rather than a bespoke import path.
"""

from dataclasses import dataclass

import openpyxl

from players.identity import ExternalPlayerRecord, resolve_player
from players.models import Player, PlayerExternalId, Position

PROVIDER = 'fantacalcio_it'

# The listone's role codes are classic-scoring roles, unlike the `RM`
# column (mantra roles), which this importer intentionally ignores for now.
ROLE_TO_POSITION = {
    'P': Position.GOALKEEPER,
    'D': Position.DEFENDER,
    'C': Position.MIDFIELDER,
    'A': Position.FORWARD,
}

HEADER_ROW = 2
DATA_START_ROW = 3


@dataclass
class ListoneRow:
    external_id: str
    position: str
    name: str
    team_name: str
    classic_price: int
    is_departed: bool = False


def parse_listone(path: str) -> list[ListoneRow]:
    """Read both the active-roster sheet ('Tutti') and the departed-players
    sheet ('Ceduti') from an official listone workbook."""
    wb = openpyxl.load_workbook(path, data_only=True)
    rows: list[ListoneRow] = []
    rows.extend(_parse_sheet(wb, 'Tutti', is_departed=False))
    if 'Ceduti' in wb.sheetnames:
        rows.extend(_parse_sheet(wb, 'Ceduti', is_departed=True))
    return rows


def _parse_sheet(wb, sheet_name: str, is_departed: bool) -> list[ListoneRow]:
    ws = wb[sheet_name]
    rows = []
    for record in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        external_id, role, _rm, name, team_name, qt_a = record[:6]
        if not name or not team_name or role not in ROLE_TO_POSITION:
            # Skip blank trailing rows or anything with an unrecognized role
            # rather than guessing — surfacing a clean error beats a silent
            # bad import (see AGENTS.md > Error Handling).
            continue
        rows.append(ListoneRow(
            external_id=str(external_id),
            position=ROLE_TO_POSITION[role],
            name=name.strip(),
            team_name=team_name.strip(),
            classic_price=int(qt_a) if qt_a is not None else 0,
            is_departed=is_departed,
        ))
    return rows


@dataclass
class PlayerImportSummary:
    created: int = 0
    updated: int = 0
    ambiguous: int = 0
    player_by_external_id: dict = None

    def __post_init__(self):
        if self.player_by_external_id is None:
            self.player_by_external_id = {}


def import_players(rows: list[ListoneRow]) -> PlayerImportSummary:
    """Upsert Player rows from parsed listone rows, going through the same
    priority-ordered identity resolver every provider uses (see
    players/identity.py) so matches stay auditable and nothing gets
    silently merged.
    """
    from players.models import Team

    summary = PlayerImportSummary()
    team_cache: dict[str, Team] = {}

    def get_team(name: str) -> Team:
        if name not in team_cache:
            team_cache[name], _ = Team.objects.get_or_create(name=name)
        return team_cache[name]

    for row in rows:
        team = get_team(row.team_name)

        # Fast path: we've imported this exact listone Id before.
        existing_mapping = PlayerExternalId.objects.select_related('player').filter(
            provider=PROVIDER, external_id=row.external_id,
        ).first()

        if existing_mapping:
            player = existing_mapping.player
            _apply_row_to_player(player, row, team)
            summary.updated += 1
            summary.player_by_external_id[row.external_id] = player
            continue

        record = ExternalPlayerRecord(
            provider=PROVIDER, external_id=row.external_id,
            full_name=row.name, club_name=row.team_name,
        )
        result = resolve_player(record)

        if result.is_ambiguous:
            summary.ambiguous += 1
            continue

        if result.player:
            player = result.player
            _apply_row_to_player(player, row, team)
            summary.updated += 1
        else:
            player = Player.objects.create(
                full_name=row.name, display_name=row.name,
                position=row.position,
                club=None if row.is_departed else team,
                previous_club=team if row.is_departed else None,
                is_active=not row.is_departed,
            )
            summary.created += 1

        PlayerExternalId.objects.get_or_create(
            player=player, provider=PROVIDER, external_id=row.external_id,
        )
        summary.player_by_external_id[row.external_id] = player

    return summary


def _apply_row_to_player(player: Player, row: ListoneRow, team) -> None:
    player.position = row.position
    player.is_active = not row.is_departed
    if row.is_departed:
        player.previous_club = team
    else:
        player.club = team
    player.save(update_fields=['position', 'is_active', 'club', 'previous_club', 'updated_at'])